# coding:utf-8

from os import path, sep
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.comments import Comment
from openpyxl.styles import Font

def test1():
    current_dir = path.dirname(path.abspath(__file__))
    file_path = f'{current_dir}{sep}openpyxl.xlsx'

    # 创建空表格
    wb = Workbook()

    # 读取表格
    # wb = load_workbook(file_path)

    # grab the active worksheet
    ws = wb.active
    # ws2 = wb.create_sheet("Sheet2") # insert at the end (default)
    # ws3 = wb.create_sheet("Sheet3") # insert at the end (default)
    # ws = wb.create_sheet("Mysheet", 0) # insert at first position
    # ws = wb.create_sheet("Mysheet", -1) # insert at the penultimate position


    # target = wb.copy_worksheet(ws)

    # sheet 标题
    ws.title = "Sheet Title"
    # 通过标题获取 Sheet 对象
    # ws3 = wb["Sheet Title"]
    # sheet 标题的背景色
    ws.sheet_properties.tabColor = "FF0000"

    print('\n ==============输出工作表名称================\n')
    print(wb.sheetnames)
    for sheet in wb:
        print(sheet.title, sheet.max_row, sheet.max_column)
    
    # Data can be assigned directly to cells
    ws['A1'] = 42
    ws['B1'] = 43
    ws['C1'] = 44

    print('\n================ 写数据 =========================\n')

    # Rows can also be appended
    ws.append([1, 2, 3])

    value = ws.cell(4, 4, 44)
    print("Cell 对象: ", value)

    # Python types will automatically be converted

    ws['A3'] = datetime.now()
    print('number_format: ', ws['A3'].number_format)

    special_cell = WriteOnlyCell(ws, value="hello world")
    special_cell.font = Font(name='Courier', size=12)
    special_cell.comment = Comment(text="A comment", author="Author's Name")
    ws.append([special_cell, '张三', None])

    # 在第 4 行上插入一行，第 4 行变成空行
    # ws.insert_rows(4)
    ws.delete_rows(4)

    # Save the file
    print(f'文件保存路径：{current_dir}')
    # This operation will overwrite existing files without warning.
    wb.save(file_path)

    print('\n================ 读数据 =========================\n')
    cell_range = ws['A1':'C2']
    print('获取 A1 - C2 之间的数据：')
    for cell in cell_range:
        print(cell)
    print()

    # 访问一组单元格
    # colC = ws['C']
    # col_range_C_D = ws['C:D']
    # row10 = ws[10]
    # row_range_5_10 = ws[5:10]

    
    """ for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
        for cell in row:
            print(cell) """

    # For performance reasons the Worksheet.iter_cols() method is not available in read-only mode.
    for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
        list = []
        for cell in col:
            list.append(cell.value)
        print(list)
    print()

    """ print('获取所有的行数据：')
    for cell in ws.rows:
        print(cell)
    print() """

    # For performance reasons the Worksheet.columns property is not available in read-only mode.
    """ print('获取所有的列数据：')
    for cell in ws.columns:
        print(cell)
    print() """

    cell_values = ws.values
    for row in cell_values:
        list = []
        for col in row:
            list.append(col)
        print(list)

    wb.close

if __name__ == "__main__":
    test1()